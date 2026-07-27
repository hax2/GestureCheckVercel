#!/usr/bin/env python3
"""Build the private research workbook and aggregate-only dashboard data."""

from __future__ import annotations

import argparse
import json
import math
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from scipy import stats


DIMENSIONS = [
    ("iconicity", "Iconicity"),
    ("sensorimotor_imagery", "Sensorimotor imagery"),
    ("motional_salience_gesture", "Motional salience"),
    ("emotional_salience_facial_expression", "Facial emotional salience"),
    ("gesture_complexity_fit", "Gesture–complexity fit"),
    ("cultural_familiarity", "Cultural familiarity"),
    ("enactment_potential", "Enactment potential"),
]
DIM_LABELS = dict(DIMENSIONS)
MODELS = [
    ("gemini_pro", "Gemini Pro", "all_rating_pro"),
    ("gemini_flash", "Gemini Flash", "all_rating_flash"),
    ("qwen", "Qwen 3.5 397B A17B", "qwen_qwen3.5-397b-a17b_video_fixed"),
]
MODEL_LABELS = {key: label for key, label, _ in MODELS}
LANGUAGE_LABELS = {"en": "English", "de": "German", "it": "Italian", "ja": "Japanese"}

NAVY = "172033"
INK = "1E293B"
BLUE = "2F6BFF"
TEAL = "13B8A6"
AMBER = "F4B740"
PALE = "E8EDF7"
WHITE = "FFFFFF"


def finite(value):
    if value is None:
        return None
    try:
        value = float(value)
    except (TypeError, ValueError):
        return None
    return value if math.isfinite(value) else None


def rounded(value, digits=4):
    value = finite(value)
    return round(value, digits) if value is not None else None


def ci95(values):
    values = np.asarray(values, dtype=float)
    values = values[np.isfinite(values)]
    if len(values) < 2:
        return (None, None)
    mean = values.mean()
    margin = stats.t.ppf(0.975, len(values) - 1) * stats.sem(values)
    return (float(mean - margin), float(mean + margin))


def safe_pearson(a, b):
    if len(a) < 3 or np.std(a) == 0 or np.std(b) == 0:
        return (None, None)
    result = stats.pearsonr(a, b)
    return (float(result.statistic), float(result.pvalue))


def safe_spearman(a, b):
    if len(a) < 3 or np.std(a) == 0 or np.std(b) == 0:
        return (None, None)
    result = stats.spearmanr(a, b)
    return (float(result.statistic), float(result.pvalue))


def bh_adjust(pvalues):
    indexed = [(i, p) for i, p in enumerate(pvalues) if p is not None and math.isfinite(p)]
    adjusted = [None] * len(pvalues)
    if not indexed:
        return adjusted
    ranked = sorted(indexed, key=lambda item: item[1])
    running = 1.0
    for rank_index in range(len(ranked) - 1, -1, -1):
        original_index, pvalue = ranked[rank_index]
        rank = rank_index + 1
        running = min(running, pvalue * len(ranked) / rank)
        adjusted[original_index] = min(1.0, running)
    return adjusted


def load_inputs(human_path, models_root, manifest_path):
    human = json.loads(human_path.read_text())
    manifest = json.loads(manifest_path.read_text())
    manifest_by_title = {item["title"]: item for item in manifest}
    model_data = {}
    for model_key, _, directory in MODELS:
        rows = {}
        for file_path in sorted((models_root / directory).glob("*.rating.json")):
            record = json.loads(file_path.read_text())
            rows[record["video_file"]] = record
        model_data[model_key] = rows
    return human, manifest_by_title, model_data


def human_frame(human):
    responses = human["responses"]
    first_seen = {}
    for row in responses:
        first_seen.setdefault(row["rater_key"], row["submitted_at"])
    ordered = sorted(first_seen, key=lambda key: (first_seen[key], key))
    rater_codes = {key: f"R{index:03d}" for index, key in enumerate(ordered, start=1)}
    rows = []
    for row in responses:
        out = {
            "rater_code": rater_codes[row["rater_key"]],
            "language": row["language"],
            "collection": row["collection"],
            "source": row["source"],
            "video_title": row["title"],
            "target_word": row["target_word"],
            "order_index": row["order_index"],
            "submitted_date_utc": str(row["submitted_at"])[:10],
        }
        for key, _ in DIMENSIONS:
            out[key] = int(row["ratings"][key])
        rows.append(out)
    return pd.DataFrame(rows), rater_codes


def model_frames(model_data, manifest):
    rating_rows = []
    rationale_rows = []
    for model_key, model_label, _ in MODELS:
        for title, record in sorted(model_data[model_key].items()):
            meta = manifest.get(title, {})
            coherence = record.get("coherence_check") or {}
            rating_row = {
                "model": model_label,
                "video_title": title,
                "target_word": record.get("target_word") or meta.get("target_word", ""),
                "collection": meta.get("collection", ""),
                "source": meta.get("source", ""),
                "gesture_description": record.get("brief_gesture_description", ""),
                "video_interpretable": coherence.get("is_video_interpretable"),
                "confidence": coherence.get("confidence", ""),
                "possible_ambiguities": " | ".join(coherence.get("possible_ambiguities") or []),
            }
            for dim, _ in DIMENSIONS:
                rating = record["ratings"][dim]
                rating_row[dim] = int(rating["score"])
                rationale_rows.append(
                    {
                        "model": model_label,
                        "video_title": title,
                        "target_word": rating_row["target_word"],
                        "dimension": DIM_LABELS[dim],
                        "score": int(rating["score"]),
                        "rationale": rating.get("rationale", ""),
                    }
                )
            rating_rows.append(rating_row)
    return pd.DataFrame(rating_rows), pd.DataFrame(rationale_rows)


def summarize(human_df, model_data, manifest):
    grouped = human_df.groupby("video_title")
    all_titles = sorted(set(manifest) | {title for rows in model_data.values() for title in rows})
    videos = []
    long_rows = []
    for title in all_titles:
        meta = manifest.get(title, {})
        subset = grouped.get_group(title) if title in grouped.groups else human_df.iloc[0:0]
        first_model = next((model_data[key].get(title) for key, _, _ in MODELS if title in model_data[key]), {})
        video = {
            "title": title,
            "target_word": meta.get("target_word") or first_model.get("target_word", ""),
            "collection": meta.get("collection", ""),
            "source": meta.get("source", ""),
            "human_response_count": int(len(subset)),
            "human": {},
            "human_by_language": {},
            "models": {},
        }
        for dim, label in DIMENSIONS:
            values = subset[dim].astype(float).to_numpy() if len(subset) else np.array([])
            lo, hi = ci95(values)
            hstats = {
                "n": int(len(values)),
                "mean": rounded(np.mean(values) if len(values) else None, 3),
                "sd": rounded(np.std(values, ddof=1) if len(values) > 1 else None, 3),
                "median": rounded(np.median(values) if len(values) else None, 3),
                "ci_low": rounded(lo, 3),
                "ci_high": rounded(hi, 3),
            }
            video["human"][dim] = hstats
            row = {
                "video_title": title,
                "target_word": video["target_word"],
                "collection": video["collection"],
                "source": video["source"],
                "dimension": label,
                "human_n": hstats["n"],
                "human_mean": hstats["mean"],
                "human_sd": hstats["sd"],
                "human_median": hstats["median"],
                "human_ci95_low": hstats["ci_low"],
                "human_ci95_high": hstats["ci_high"],
            }
            for model_key, model_label, _ in MODELS:
                record = model_data[model_key].get(title)
                score = int(record["ratings"][dim]["score"]) if record else None
                row[model_label] = score
                row[f"{model_label} minus human"] = rounded(score - hstats["mean"], 3) if score and hstats["mean"] else None
            long_rows.append(row)
        for language, language_subset in subset.groupby("language"):
            video["human_by_language"][language] = {}
            for dim, _ in DIMENSIONS:
                values = language_subset[dim].astype(float).to_numpy()
                video["human_by_language"][language][dim] = {
                    "n": int(len(values)),
                    "mean": rounded(np.mean(values), 3),
                    "sd": rounded(np.std(values, ddof=1) if len(values) > 1 else None, 3),
                }
        for model_key, _, _ in MODELS:
            record = model_data[model_key].get(title)
            if not record:
                continue
            video["models"][model_key] = {
                "description": record.get("brief_gesture_description", ""),
                "scores": {dim: int(record["ratings"][dim]["score"]) for dim, _ in DIMENSIONS},
                "rationales": {dim: record["ratings"][dim].get("rationale", "") for dim, _ in DIMENSIONS},
                "coherence": record.get("coherence_check") or {},
            }
        videos.append(video)
    return videos, pd.DataFrame(long_rows)


def overall_statistics(videos):
    rows = []
    for model_key, model_label, _ in MODELS:
        for dim, label in DIMENSIONS:
            pairs = [
                (video["human"][dim]["mean"], video["models"].get(model_key, {}).get("scores", {}).get(dim))
                for video in videos
            ]
            pairs = [(h, m) for h, m in pairs if h is not None and m is not None]
            human = np.array([pair[0] for pair in pairs], dtype=float)
            model = np.array([pair[1] for pair in pairs], dtype=float)
            diff = model - human
            pearson_r, pearson_p = safe_pearson(human, model)
            spearman_rho, spearman_p = safe_spearman(human, model)
            ttest = stats.ttest_rel(model, human) if len(diff) > 1 else None
            try:
                wilcoxon = stats.wilcoxon(diff) if len(diff) > 1 and np.any(diff != 0) else None
            except ValueError:
                wilcoxon = None
            bias_lo, bias_hi = ci95(diff)
            rows.append(
                {
                    "model_key": model_key,
                    "model": model_label,
                    "dimension_key": dim,
                    "dimension": label,
                    "paired_videos": len(pairs),
                    "human_video_mean": rounded(np.mean(human)),
                    "model_mean": rounded(np.mean(model)),
                    "mean_bias_model_minus_human": rounded(np.mean(diff)),
                    "bias_ci95_low": rounded(bias_lo),
                    "bias_ci95_high": rounded(bias_hi),
                    "mae": rounded(np.mean(np.abs(diff))),
                    "rmse": rounded(np.sqrt(np.mean(diff**2))),
                    "pearson_r": rounded(pearson_r),
                    "pearson_p": rounded(pearson_p, 6),
                    "spearman_rho": rounded(spearman_rho),
                    "spearman_p": rounded(spearman_p, 6),
                    "paired_t": rounded(ttest.statistic if ttest else None),
                    "paired_t_p": rounded(ttest.pvalue if ttest else None, 6),
                    "cohens_dz": rounded(np.mean(diff) / np.std(diff, ddof=1) if len(diff) > 1 and np.std(diff, ddof=1) else None),
                    "wilcoxon_w": rounded(wilcoxon.statistic if wilcoxon else None),
                    "wilcoxon_p": rounded(wilcoxon.pvalue if wilcoxon else None, 6),
                }
            )
    adjusted = bh_adjust([row["paired_t_p"] for row in rows])
    for row, pvalue in zip(rows, adjusted):
        row["paired_t_p_bh"] = rounded(pvalue, 6)
    return rows


def model_agreement(videos):
    rows = []
    for left_index, (left_key, left_label, _) in enumerate(MODELS):
        for right_key, right_label, _ in MODELS[left_index + 1 :]:
            for dim, label in DIMENSIONS:
                pairs = []
                for video in videos:
                    left = video["models"].get(left_key, {}).get("scores", {}).get(dim)
                    right = video["models"].get(right_key, {}).get("scores", {}).get(dim)
                    if left is not None and right is not None:
                        pairs.append((left, right))
                a = np.array([pair[0] for pair in pairs], dtype=float)
                b = np.array([pair[1] for pair in pairs], dtype=float)
                pearson_r, pearson_p = safe_pearson(a, b)
                spearman_rho, spearman_p = safe_spearman(a, b)
                rows.append(
                    {
                        "model_a": left_label,
                        "model_b": right_label,
                        "dimension": label,
                        "videos": len(pairs),
                        "mean_a": rounded(np.mean(a)),
                        "mean_b": rounded(np.mean(b)),
                        "mean_absolute_difference": rounded(np.mean(np.abs(a - b))),
                        "exact_agreement_percent": rounded(np.mean(a == b) * 100, 2),
                        "within_one_point_percent": rounded(np.mean(np.abs(a - b) <= 1) * 100, 2),
                        "pearson_r": rounded(pearson_r),
                        "pearson_p": rounded(pearson_p, 6),
                        "spearman_rho": rounded(spearman_rho),
                        "spearman_p": rounded(spearman_p, 6),
                    }
                )
    return rows


def build_comments(overall, human_df):
    comments = []
    by_model = defaultdict(list)
    for row in overall:
        by_model[row["model"]].append(row)
    mean_maes = {model: np.mean([row["mae"] for row in rows]) for model, rows in by_model.items()}
    best_model = min(mean_maes, key=mean_maes.get)
    comments.append(
        {
            "title": "Closest model overall",
            "text": f"{best_model} has the lowest mean dimension-level MAE ({mean_maes[best_model]:.2f}) against video-level human means.",
        }
    )
    for model, rows in by_model.items():
        strongest = max((row for row in rows if row["pearson_r"] is not None), key=lambda row: row["pearson_r"])
        largest_bias = max(rows, key=lambda row: abs(row["mean_bias_model_minus_human"]))
        direction = "higher" if largest_bias["mean_bias_model_minus_human"] > 0 else "lower"
        comments.append(
            {
                "title": model,
                "text": (
                    f"Strongest linear alignment is on {strongest['dimension']} (r={strongest['pearson_r']:.2f}). "
                    f"The largest mean offset is {largest_bias['dimension']}, where scores are "
                    f"{abs(largest_bias['mean_bias_model_minus_human']):.2f} points {direction} than human means."
                ),
            }
        )
    language_counts = human_df["language"].value_counts()
    dominant = language_counts.index[0]
    comments.append(
        {
            "title": "Sampling caution",
            "text": (
                f"{LANGUAGE_LABELS.get(dominant, dominant)}-interface ratings make up "
                f"{language_counts.iloc[0] / len(human_df):.1%} of complete responses. "
                "The pooled human benchmark should not be interpreted as culturally balanced."
            ),
        }
    )
    comments.append(
        {
            "title": "Inference boundary",
            "text": (
                "Model–human tests are exploratory item-level comparisons. A model score is one generated judgment, "
                "while each human value is a mean from a varying number of raters; correlation does not establish interchangeability."
            ),
        }
    )
    return comments


def aggregate_demographics(human, rater_codes):
    rows = []
    completed_keys = set(rater_codes)
    records = [p for p in human["participants"] if p["rater_key"] in completed_keys]
    for field in [
        "gender",
        "native_language",
        "country_of_origin",
        "education",
        "handedness",
        "gesture_culture_familiarity",
        "geo_country",
    ]:
        values = [str((row.get("demographics") or {}).get(field, "")).strip() for row in records]
        values = [value for value in values if value]
        for value, count in sorted(Counter(values).items(), key=lambda item: (-item[1], item[0])):
            rows.append(
                {
                    "field": field,
                    "category": value,
                    "count": count,
                    "percent_of_nonmissing": round(count / len(values) * 100, 2) if values else None,
                    "nonmissing_n": len(values),
                }
            )
    ages = []
    for row in records:
        try:
            ages.append(float((row.get("demographics") or {}).get("age")))
        except (TypeError, ValueError):
            pass
    if ages:
        for category, value in [
            ("mean", np.mean(ages)),
            ("sd", np.std(ages, ddof=1) if len(ages) > 1 else None),
            ("median", np.median(ages)),
            ("minimum", np.min(ages)),
            ("maximum", np.max(ages)),
        ]:
            rows.append(
                {
                    "field": "age_summary",
                    "category": category,
                    "count": rounded(value, 2),
                    "percent_of_nonmissing": None,
                    "nonmissing_n": len(ages),
                }
            )
    return pd.DataFrame(rows)


def add_table(ws, name):
    if ws.max_row < 2 or ws.max_column < 1:
        return
    table = Table(displayName=name, ref=f"A1:{get_column_letter(ws.max_column)}{ws.max_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(table)


def style_sheet(ws, freeze="A2", widths=None):
    ws.freeze_panes = freeze
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 34
    for column_index in range(1, ws.max_column + 1):
        header = str(ws.cell(1, column_index).value or "")
        width = widths.get(header) if widths else None
        if width is None:
            width = min(max(len(header) + 2, 11), 24)
        ws.column_dimensions[get_column_letter(column_index)].width = width


def dataframe_sheet(wb, title, frame, table_name, widths=None):
    ws = wb.create_sheet(title)
    ws.append(list(frame.columns))
    for row in frame.itertuples(index=False, name=None):
        ws.append([None if pd.isna(value) else value for value in row])
    style_sheet(ws, widths=widths)
    add_table(ws, table_name)
    return ws


def build_workbook(output_path, human_df, model_df, rationales_df, video_df, overall, agreement, demographics, comments, metadata):
    wb = Workbook()
    readme = wb.active
    readme.title = "Read Me"
    readme.sheet_view.showGridLines = False
    readme.column_dimensions["A"].width = 28
    readme.column_dimensions["B"].width = 112
    readme["A1"] = "Gesture ratings analysis"
    readme["A1"].font = Font(size=22, bold=True, color=WHITE)
    readme["A1"].fill = PatternFill("solid", fgColor=NAVY)
    readme.merge_cells("A1:B1")
    rows = [
        ("Purpose", "Human–VLM comparison across seven 1–5 gesture-rating dimensions."),
        ("Snapshot", metadata["generated_at"]),
        ("Human data", f"{metadata['human_responses']} complete responses, {metadata['human_raters']} pseudonymized raters, {metadata['human_videos']} videos."),
        ("VLM data", "149 videos for each of Gemini Pro, Gemini Flash, and corrected Qwen 3.5 397B A17B."),
        ("Privacy", "Raw database identifiers, exact submission times, raw payloads, and participant-level demographics are excluded. Rater codes are study-local pseudonyms."),
        ("Primary comparison", "Each VLM score is paired with the human mean for the same video and dimension. Videos—not individual ratings—receive equal weight."),
        ("Inference", "Pearson and Spearman correlation measure item ordering; MAE/RMSE measure distance; bias tests whether model scores are systematically higher/lower."),
        ("Multiple testing", "Paired t-test p-values are Benjamini–Hochberg adjusted across the 21 model × dimension tests."),
        ("Important limitation", "Human n varies by video (5–17), language composition is uneven, and VLM values are single generated outputs."),
        ("Missing pairing", "87_Cooking.mp4 has VLM ratings but no complete human rating in this snapshot."),
    ]
    for row_index, (label, value) in enumerate(rows, start=3):
        readme.cell(row_index, 1, label).font = Font(bold=True, color=INK)
        readme.cell(row_index, 2, value).alignment = Alignment(wrap_text=True, vertical="top")

    findings = wb.create_sheet("Key Findings")
    findings.sheet_view.showGridLines = False
    findings.column_dimensions["A"].width = 30
    findings.column_dimensions["B"].width = 112
    findings.append(["Finding", "Interpretation"])
    for comment in comments:
        findings.append([comment["title"], comment["text"]])
    style_sheet(findings)
    for row in findings.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        row[0].font = Font(bold=True, color=INK)
        findings.row_dimensions[row[0].row].height = 48

    human_ws = dataframe_sheet(
        wb,
        "Human Ratings",
        human_df,
        "HumanRatingsTable",
        widths={"video_title": 28, "target_word": 22, "submitted_date_utc": 21},
    )
    for col in range(9, 16):
        human_ws.conditional_formatting.add(
            f"{get_column_letter(col)}2:{get_column_letter(col)}{human_ws.max_row}",
            ColorScaleRule(start_type="num", start_value=1, start_color="F8696B", mid_type="num", mid_value=3, mid_color="FFEB84", end_type="num", end_value=5, end_color="63BE7B"),
        )

    dataframe_sheet(
        wb,
        "VLM Ratings",
        model_df,
        "VLMRatingsTable",
        widths={"video_title": 28, "target_word": 22, "gesture_description": 70, "possible_ambiguities": 70},
    )
    dataframe_sheet(
        wb,
        "VLM Rationales",
        rationales_df,
        "VLMRationalesTable",
        widths={"video_title": 28, "target_word": 22, "dimension": 30, "rationale": 100},
    )
    dataframe_sheet(
        wb,
        "Video Dimension Summary",
        video_df,
        "VideoSummaryTable",
        widths={"video_title": 28, "target_word": 22, "dimension": 31},
    )
    overall_df = pd.DataFrame(overall).drop(columns=["model_key", "dimension_key"])
    overall_ws = dataframe_sheet(wb, "Overall Statistics", overall_df, "OverallStatsTable", widths={"dimension": 31})
    for column_index, column_name in enumerate(overall_df.columns, start=1):
        if column_name in {"mae", "rmse", "pearson_r", "spearman_rho", "mean_bias_model_minus_human"}:
            overall_ws.conditional_formatting.add(
                f"{get_column_letter(column_index)}2:{get_column_letter(column_index)}{overall_ws.max_row}",
                ColorScaleRule(start_type="min", start_color="F8696B", mid_type="percentile", mid_value=50, mid_color="FFEB84", end_type="max", end_color="63BE7B"),
            )

    agreement_df = pd.DataFrame(agreement)
    dataframe_sheet(wb, "Model Agreement", agreement_df, "ModelAgreementTable", widths={"dimension": 31})
    dataframe_sheet(wb, "Demographics Aggregate", demographics, "DemographicsTable", widths={"field": 34, "category": 32})

    coverage_rows = []
    for title, group in human_df.groupby("video_title"):
        coverage_rows.append(
            {
                "video_title": title,
                "target_word": group.iloc[0]["target_word"],
                "human_responses": len(group),
                "distinct_raters": group["rater_code"].nunique(),
                **{f"{lang}_responses": int((group["language"] == lang).sum()) for lang in LANGUAGE_LABELS},
            }
        )
    coverage = pd.DataFrame(coverage_rows).sort_values("video_title")
    dataframe_sheet(wb, "Coverage", coverage, "CoverageTable", widths={"video_title": 28, "target_word": 22})

    chart_ws = wb.create_sheet("Charts")
    chart_ws.sheet_view.showGridLines = False
    chart_ws["A1"] = "Mean rating by dimension"
    chart_ws["A1"].font = Font(size=18, bold=True, color=INK)
    chart_ws.append(["Dimension", "Human", *[label for _, label, _ in MODELS]])
    for dim, label in DIMENSIONS:
        rows = [row for row in overall if row["dimension_key"] == dim]
        chart_ws.append([label, rows[0]["human_video_mean"], *[next(row["model_mean"] for row in rows if row["model_key"] == key) for key, _, _ in MODELS]])
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "Human and VLM means"
    chart.y_axis.title = "Mean score (1–5)"
    chart.y_axis.scaling.min = 1
    chart.y_axis.scaling.max = 5
    chart.x_axis.title = "Dimension"
    data = Reference(chart_ws, min_col=2, max_col=5, min_row=2, max_row=9)
    categories = Reference(chart_ws, min_col=1, min_row=3, max_row=9)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    chart.height = 11
    chart.width = 24
    chart_ws.add_chart(chart, "A12")
    chart_ws.column_dimensions["A"].width = 32
    for col in "BCDE":
        chart_ws.column_dimensions[col].width = 18

    methodology = wb.create_sheet("Methods & Notes")
    methodology.sheet_view.showGridLines = False
    methodology.column_dimensions["A"].width = 30
    methodology.column_dimensions["B"].width = 112
    methods = [
        ("Human unit", "One completed response for one video, containing all seven integer ratings."),
        ("Human benchmark", "Arithmetic mean per video × dimension across all complete responses."),
        ("Model unit", "One integer score per video × model × dimension, from the supplied corrected output directories."),
        ("Bias", "Mean(model score − human video mean). Positive values mean the model rates higher."),
        ("MAE", "Mean absolute model–human difference in scale points; lower is closer."),
        ("RMSE", "Root mean squared difference; penalizes larger discrepancies more strongly than MAE."),
        ("Pearson r", "Linear association across videos. High r can coexist with substantial mean bias."),
        ("Spearman rho", "Rank-order association across videos; less sensitive to scale spacing."),
        ("Paired t-test", "Exploratory test of whether mean paired difference is zero across videos."),
        ("Cohen's dz", "Standardized paired mean difference."),
        ("Wilcoxon", "Exploratory signed-rank test of paired differences."),
        ("BH correction", "Controls false-discovery rate across all 21 paired t-tests in this workbook."),
        ("Language", "Interface language is retained in Human Ratings and dashboard aggregates; it is not necessarily the rater's native language."),
        ("Privacy", "No raw participant/session/response IDs, raw payloads, exact times, or row-level demographics are exported."),
        ("Reproducibility", "Run scripts/export_human_ratings.mjs with DATABASE_URL, then scripts/build_rating_analysis.py."),
    ]
    methodology.append(["Topic", "Detail"])
    for item in methods:
        methodology.append(item)
    style_sheet(methodology)
    for row in methodology.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
        methodology.row_dimensions[row[0].row].height = 42

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def dashboard_payload(videos, overall, agreement, comments, human_df, generated_at):
    for video in videos:
        video["languages"] = sorted(video.pop("human_by_language").keys())
    return {
        "metadata": {
            "generated_at": generated_at,
            "human_responses": int(len(human_df)),
            "human_raters": int(human_df["rater_code"].nunique()),
            "human_videos": int(human_df["video_title"].nunique()),
            "model_videos": 149,
            "human_ratings_total": int(len(human_df) * len(DIMENSIONS)),
            "models": [{"key": key, "label": label} for key, label, _ in MODELS],
            "languages": [{"key": key, "label": label, "responses": int((human_df["language"] == key).sum())} for key, label in LANGUAGE_LABELS.items()],
        },
        "dimensions": [{"key": key, "label": label} for key, label in DIMENSIONS],
        "overall": overall,
        "model_agreement": agreement,
        "comments": comments,
        "videos": videos,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--human", type=Path, default=Path("analysis/source/human_ratings.json"))
    parser.add_argument("--models-root", type=Path, default=Path("/home/Shodan/Projects/GestureCheck/results"))
    parser.add_argument("--manifest", type=Path, default=Path("public/all_rating_videos.json"))
    parser.add_argument("--workbook", type=Path, default=Path("analysis/gesture_ratings_complete_analysis.xlsx"))
    parser.add_argument("--dashboard-data", type=Path, default=Path("public/research-insights-7f3c9a/data.json"))
    args = parser.parse_args()

    human, manifest, model_data = load_inputs(args.human, args.models_root, args.manifest)
    human_df, rater_codes = human_frame(human)
    model_df, rationale_df = model_frames(model_data, manifest)
    videos, video_df = summarize(human_df, model_data, manifest)
    overall = overall_statistics(videos)
    agreement = model_agreement(videos)
    comments = build_comments(overall, human_df)
    demographics = aggregate_demographics(human, rater_codes)
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    metadata = {
        "generated_at": generated_at,
        "human_responses": len(human_df),
        "human_raters": human_df["rater_code"].nunique(),
        "human_videos": human_df["video_title"].nunique(),
    }
    build_workbook(
        args.workbook,
        human_df,
        model_df,
        rationale_df,
        video_df,
        overall,
        agreement,
        demographics,
        comments,
        metadata,
    )
    args.dashboard_data.parent.mkdir(parents=True, exist_ok=True)
    args.dashboard_data.write_text(
        json.dumps(dashboard_payload(videos, overall, agreement, comments, human_df, generated_at), ensure_ascii=False, separators=(",", ":"))
        + "\n"
    )
    print(
        json.dumps(
            {
                "workbook": str(args.workbook),
                "dashboard_data": str(args.dashboard_data),
                "human_responses": len(human_df),
                "paired_videos": sum(video["human_response_count"] > 0 for video in videos),
                "model_videos": len(videos),
            }
        )
    )


if __name__ == "__main__":
    main()
